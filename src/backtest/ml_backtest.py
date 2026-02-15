"""
ML 예측 기반 Walk-Forward 백테스트 시스템

이 모듈은 regressor.py가 생성한 예측 캐시를 사용하여 백테스트를 수행합니다.
모델 학습/예측은 regressor.py에서만 수행되며, 이 모듈은 수익률 계산에만 집중합니다.

핵심 원칙:
1. 일원화: regressor.py의 예측 캐시를 반드시 사용 (독립 학습 금지)
2. Filing Date(공시일)를 엄격히 준수
3. 미래 데이터는 절대 사용하지 않음

작성자: Quant Trading Team
날짜: 2025-11-20
"""

import logging
import joblib
import pandas as pd
import numpy as np
from datetime import datetime
from dateutil.relativedelta import relativedelta
from pathlib import Path
from typing import Dict, Any, Tuple

# ✨ 리팩토링: 통합 데이터 스키마 및 전처리 import
# DataProcessor가 모든 스케일링(RobustScaler, StandardScaler)을 처리합니다
from src.constants.data_schema import DataSchema
from src.training.data_processor import DataProcessor


class MLBacktest:
    """
    regressor.py의 예측 캐시를 사용한 Walk-Forward 백테스트

    백테스트 플로우:
    ---------------
    1. regressor.py가 생성한 predictions cache 로드 (필수)
    2. 리밸런싱 날짜 목록 생성 (예: 2023-03-13, 2023-06-13, ...)
    3. 각 리밸런싱 날짜마다:
       a. 캐시에서 예측 결과 로드
       b. 상위 K개 종목 선택
       c. 실제 수익률 계산
    4. 전체 성과 리포트 생성

    일원화 원칙:
    -----------
    - 모델 학습/예측은 regressor.py에서만 수행
    - ml_backtest.py는 예측 캐시를 소비하여 수익률만 계산
    - 캐시 없이는 실행 불가 (일원화 강제)

    Parameters:
    ----------
    config : Dict[str, Any]
        설정 딕셔너리 (conf.yaml에서 로드)
    main_ctx : MainContext
        메인 컨텍스트 (root_path 등 포함)
    rebalance_period : int
        리밸런싱 주기 (개월), 기본값 3
    top_k : int
        선택할 종목 수, 기본값 20
    """

    def __init__(
        self,
        config: Dict[str, Any],
        main_ctx: Any,
        rebalance_period: int = 3,
        top_k: int = 20,
    ):
        self.config = config
        self.main_ctx = main_ctx
        self.rebalance_period = rebalance_period
        self.top_k = top_k

        # 로깅 설정
        self.logger = logging.getLogger('MLBacktest')

        # 경로 설정
        self.data_path = Path(main_ctx.root_path) / 'processed' / 'ml_data' / 'per_year'

        # 결과 저장용
        self.backtest_results = []
        self.detailed_results = []  # 각 종목별 상세 거래 내역
        self.predictions_history = []

        # 예측 캐시 로드 (regressor.py가 생성한 캐시 필수)
        eval_config = config.get('EVALUATION', {})
        cache_file = eval_config.get('PREDICTIONS_CACHE_FILE', 'regressor_predictions.pkl')
        cache_path = Path(main_ctx.root_path) / 'MODELS' / cache_file

        if cache_path.exists():
            self.predictions_cache = joblib.load(cache_path)
            self.logger.info(f"✅ Loaded predictions cache: {cache_path}")
            self.logger.info(f"   Available periods: {len(self.predictions_cache)}")
            self.logger.info(f"   Dates: {list(self.predictions_cache.keys())}")
        else:
            raise FileNotFoundError(
                f"\n{'='*60}\n"
                f"❌ Predictions cache not found!\n"
                f"   Path: {cache_path}\n"
                f"\n"
                f"   regressor.py를 먼저 실행하여 예측 캐시를 생성하세요.\n"
                f"   (일원화 원칙: 모델 학습/예측은 regressor.py에서만 수행)\n"
                f"{'='*60}"
            )

        # 거래 비용 설정
        backtest_config = config.get('BACKTEST', {})
        trading_costs = backtest_config.get('TRADING_COSTS', {})
        self.trading_costs_enabled = trading_costs.get('ENABLED', 'N') == 'Y'
        self.commission = trading_costs.get('COMMISSION', 0.001)  # 0.1% 기본값
        self.slippage = trading_costs.get('SLIPPAGE', 0.001)      # 0.1% 기본값

        if self.trading_costs_enabled:
            total_cost = 2 * (self.commission + self.slippage)
            self.logger.info(f"💰 Trading costs enabled: commission={self.commission*100:.2f}%, "
                           f"slippage={self.slippage*100:.2f}%, total={total_cost*100:.2f}%/trade")
        else:
            self.logger.info("💰 Trading costs disabled (pure returns)")

        # 상장폐지 데이터 로드 (검증용)
        self.delisted_data = self._load_delisted_data()

    def _load_delisted_data(self) -> pd.DataFrame:
        """
        상장폐지 종목 데이터를 로드합니다.

        FMP의 delisted_companies 데이터를 로드하여 상장폐지 여부 검증에 사용합니다.
        데이터가 없으면 빈 DataFrame을 반환합니다 (검증 스킵).
        """
        delisted_dir = Path(self.main_ctx.root_path) / 'fmp_raw' / 'delisted_companies'

        if not delisted_dir.exists():
            self.logger.info("📋 No delisted_companies data found - delisting verification disabled")
            return pd.DataFrame()

        try:
            all_delisted = []
            for file in delisted_dir.glob('*.parquet'):
                df = pd.read_parquet(file)
                if not df.empty:
                    all_delisted.append(df)

            if not all_delisted:
                self.logger.info("📋 delisted_companies directory is empty - delisting verification disabled")
                return pd.DataFrame()

            delisted_df = pd.concat(all_delisted, ignore_index=True)

            # 필요한 컬럼만 유지합니다
            required_cols = ['symbol', 'delistedDate']
            available_cols = [c for c in required_cols if c in delisted_df.columns]

            if 'symbol' not in available_cols:
                self.logger.warning("⚠️ delisted_companies data missing 'symbol' column")
                return pd.DataFrame()

            delisted_df = delisted_df[available_cols].drop_duplicates(subset=['symbol'])

            if 'delistedDate' in delisted_df.columns:
                delisted_df['delistedDate'] = pd.to_datetime(delisted_df['delistedDate'], errors='coerce')

            self.logger.info(f"📋 Loaded {len(delisted_df)} delisted companies for verification")
            return delisted_df

        except Exception as e:
            self.logger.warning(f"⚠️ Error loading delisted data: {e}")
            return pd.DataFrame()

    def _is_truly_delisted(self, symbol: str, check_date: pd.Timestamp) -> tuple:
        """
        종목이 특정 날짜에 실제로 상장폐지되었는지 검증합니다.

        Returns:
        -------
        tuple: (is_delisted: bool, delisted_date: str or None, status: str)
            - is_delisted: 상장폐지 여부
            - delisted_date: 상장폐지 날짜 (있으면)
            - status: 'confirmed_delisted', 'data_missing', 'active'
        """
        if self.delisted_data.empty:
            # 검증 데이터 없음 - 가격 없으면 data_missing으로 처리합니다
            return (False, None, 'unverified')

        symbol_data = self.delisted_data[self.delisted_data['symbol'] == symbol]

        if symbol_data.empty:
            # 상장폐지 리스트에 없음 - 아직 상장 중이거나 데이터 누락입니다
            return (False, None, 'not_in_delisted_list')

        # 상장폐지 리스트에 있음
        if 'delistedDate' in symbol_data.columns:
            delisted_date = symbol_data.iloc[0]['delistedDate']
            if pd.notna(delisted_date):
                # 상장폐지 날짜가 체크 날짜 이전인지 확인합니다
                if delisted_date <= check_date:
                    return (True, delisted_date.strftime('%Y-%m-%d'), 'confirmed_delisted')
                else:
                    # 아직 상장폐지 전입니다
                    return (False, delisted_date.strftime('%Y-%m-%d'), 'will_be_delisted')

        # 상장폐지 날짜 정보 없음
        return (True, None, 'delisted_no_date')

    def _select_top_k(self, predictions: pd.DataFrame) -> pd.DataFrame:
        """
        상위 K개 종목을 선택합니다.

        Parameters:
        ----------
        predictions : pd.DataFrame
            예측 결과 (symbol, sector, ml_score 등 포함)

        Returns:
        -------
        pd.DataFrame
            선택된 종목 정보 (symbol, sector 등 포함)
        """
        # ml_score 기준으로 정렬
        sorted_df = predictions.sort_values('ml_score', ascending=False)

        # 상위 K개 선택 (symbol과 sector 모두 포함합니다)
        top_k_df = sorted_df.head(self.top_k).copy()

        # sector 값 보장 (빈 sector 이슈 수정)
        if 'sector' not in top_k_df.columns:
            top_k_df['sector'] = 'Unknown'
        else:
            top_k_df['sector'] = top_k_df['sector'].fillna('Unknown')

        # SECTOR_CATEGORIZATION 설정을 사용하여 category 컬럼 추가
        top_k_df = DataProcessor.map_sectors_to_categories(
            top_k_df,
            self.config,
            sector_column='sector',
            logger=None  # 예측 시에는 로깅하지 않음
        )
        # 명확성을 위해 sector_category를 category로 이름 변경
        if 'sector_category' in top_k_df.columns:
            top_k_df['category'] = top_k_df['sector_category']
            top_k_df.drop(columns=['sector_category'], inplace=True, errors='ignore')
        else:
            top_k_df['category'] = top_k_df['sector']  # Fallback: sector를 category로 사용

        return top_k_df[['symbol', 'sector', 'category']].copy()

    def _calculate_period_return(
        self,
        selected_stocks: pd.DataFrame,
        buy_date: datetime,
        sell_date: datetime,
        price_table: pd.DataFrame
    ) -> dict:
        """
        기간 수익률을 계산합니다 (상세 정보 포함).

        Parameters:
        ----------
        selected_stocks : pd.DataFrame
            선택된 종목 정보 (symbol, sector 컬럼 포함)
        buy_date : datetime
            매수 날짜
        sell_date : datetime
            매도 날짜
        price_table : pd.DataFrame
            가격 데이터

        Returns:
        -------
        dict
            {
                'avg_return': float,  # 평균 수익률
                'details': list[dict]  # 각 종목별 상세 정보
            }
        """
        # 실제 거래일 찾기 (주말/휴일 처리)
        # 일원화: DataProcessor.get_trade_date() 사용 (make_mldata.py, regressor.py 공통)
        actual_buy_date = DataProcessor.get_trade_date(pd.Timestamp(buy_date), price_table)
        actual_sell_date = DataProcessor.get_trade_date(pd.Timestamp(sell_date), price_table)

        if actual_buy_date is None or actual_sell_date is None:
            self.logger.warning(
                f"Trading date not found: buy={buy_date.date()}, sell={sell_date.date()}"
            )
            return {'avg_return': 0.0, 'details': [], 'actual_buy_date': None, 'actual_sell_date': None}

        returns = []
        details = []

        # 섹터 + 카테고리 정보 포함하여 반복
        for _, stock in selected_stocks.iterrows():
            symbol = stock['symbol']
            sector = stock.get('sector', 'Unknown')  # sector가 없으면 'Unknown' 사용
            category = stock.get('category', sector)  # category가 없으면 sector를 사용

            symbol_prices = price_table[price_table['symbol'] == symbol]

            # 매수 가격 조회 (실제 거래일)
            buy_price_rows = symbol_prices[symbol_prices['date'] == actual_buy_date]
            if buy_price_rows.empty:
                # 매수일에 가격 없음 = 데이터 오류 (거래 불가능합니다)
                self.logger.warning(f"   ⚠️  {symbol}: No price at buy date {actual_buy_date.date()} - skipping")
                continue
            buy_price = buy_price_rows.iloc[0]['close']

            # 매도 가격 조회 (실제 거래일)
            sell_price_rows = symbol_prices[symbol_prices['date'] == actual_sell_date]
            if sell_price_rows.empty:
                # 가격 데이터 없음 - 상장폐지 여부 검증
                is_delisted_verified, delisted_date, status = self._is_truly_delisted(
                    symbol, actual_sell_date
                )

                # 디버그 로그: 가격이 없는 종목의 상세 정보를 기록합니다
                last_price_date = symbol_prices['date'].max() if not symbol_prices.empty else None
                self.logger.warning(
                    f"   🔍 {symbol}: No price at {actual_sell_date.date()} | "
                    f"Last price: {last_price_date.date() if last_price_date else 'N/A'} | "
                    f"Delisting status: {status}"
                    + (f" (delisted: {delisted_date})" if delisted_date else "")
                )

                if is_delisted_verified or status in ('confirmed_delisted', 'delisted_no_date'):
                    # 확인된 상장폐지: -100% 수익률 처리
                    gross_ret = -1.0
                    net_ret = -1.0
                    sell_price = 0.0
                    is_delisted = True
                    delisted_status = 'confirmed'
                    trading_cost = 0.0
                    self.logger.warning(f"      → CONFIRMED DELISTED: -100% return")
                else:
                    # 상장폐지 미확인: 데이터 누락으로 처리합니다 (스킵하거나 마지막 가격 사용)
                    # 옵션 1: 스킵 (보수적)
                    # 옵션 2: 마지막 가격 사용 (낙관적)
                    if not symbol_prices.empty:
                        # 마지막 가격 사용 (가장 최근 거래 가격)
                        last_price_row = symbol_prices.sort_values('date').iloc[-1]
                        sell_price = last_price_row['close']
                        gross_ret = (sell_price - buy_price) / buy_price
                        net_ret = gross_ret  # 거래 비용은 별도 계산
                        is_delisted = False
                        delisted_status = 'data_missing_used_last_price'
                        trading_cost = 0.0
                        self.logger.warning(
                            f"      → DATA MISSING: Using last available price ${sell_price:.2f} "
                            f"from {last_price_row['date'].date()} (return: {gross_ret*100:.2f}%)"
                        )
                    else:
                        # 가격 데이터가 전혀 없음 - 스킵
                        self.logger.warning(f"      → NO PRICE DATA: Skipping {symbol}")
                        continue
            else:
                sell_price = sell_price_rows.iloc[0]['close']

                # 순수 수익률 (거래 비용 미반영)
                gross_ret = (sell_price - buy_price) / buy_price

                # 거래 비용 반영
                if self.trading_costs_enabled:
                    # 슬리피지: 매수 시 높게, 매도 시 낮게
                    effective_buy_price = buy_price * (1 + self.slippage)
                    effective_sell_price = sell_price * (1 - self.slippage)

                    # 슬리피지 적용 후 수익률
                    ret_after_slippage = (effective_sell_price - effective_buy_price) / effective_buy_price

                    # 거래 수수료 (매수 + 매도)
                    net_ret = ret_after_slippage - 2 * self.commission
                    trading_cost = gross_ret - net_ret
                else:
                    net_ret = gross_ret
                    trading_cost = 0.0

                is_delisted = False
                delisted_status = 'active'  # 정상 거래

            # 수익률 리스트에 추가 (상장폐지 포함) - 순수익률 사용
            returns.append(net_ret)

            # 상세 정보 저장 (섹터 + 카테고리 정보 + 상장폐지 여부 + 거래 비용)
            details.append({
                'symbol': symbol,
                'sector': sector,      # ✅ 섹터 정보 추가
                'category': category,  # ✅ 카테고리 정보 추가
                'buy_price': buy_price,
                'sell_price': sell_price,
                'gross_return': gross_ret,         # 순수 수익률
                'trading_cost': trading_cost,       # 거래 비용
                'return': net_ret,                  # 순수익률 (거래 비용 차감)
                'return_pct': net_ret * 100,
                'delisted': is_delisted,           # 상장폐지 여부 (boolean)
                'delisted_status': delisted_status  # ✅ 상세 상태 (confirmed/data_missing_used_last_price/active)
            })

        if not returns:
            return {'avg_return': 0.0, 'details': [], 'actual_buy_date': actual_buy_date, 'actual_sell_date': actual_sell_date}

        return {
            'avg_return': np.mean(returns),
            'details': details,
            'actual_buy_date': actual_buy_date,
            'actual_sell_date': actual_sell_date
        }

    def _calculate_benchmark_returns(
        self,
        start_date: datetime,
        end_date: datetime,
        price_table: pd.DataFrame
    ) -> pd.DataFrame:
        """
        벤치마크 Buy-and-Hold 수익률 계산

        ETF와 주식 데이터를 분리하여 처리합니다:
        - 주식: 기존 price_table 사용
        - ETF: etf_price.parquet 별도 로드 (ETFDataLoader 사용)

        이를 통해 ETF가 ML 학습 데이터에 오염되는 것을 방지합니다.

        Parameters:
        ----------
        start_date : datetime
            백테스트 시작 날짜
        end_date : datetime
            백테스트 종료 날짜
        price_table : pd.DataFrame
            가격 데이터 (주식 only)

        Returns:
        -------
        pd.DataFrame
            벤치마크 결과 (symbol, return, sharpe, mdd 등)
        """
        benchmark_config = self.config.get('BENCHMARK', {})

        if not benchmark_config.get('ENABLED', 'N') == 'Y':
            self.logger.info("📊 Benchmark comparison disabled (BENCHMARK.ENABLED=N)")
            return pd.DataFrame()

        benchmark_symbols = benchmark_config.get('SYMBOLS', [])
        if not benchmark_symbols:
            self.logger.warning("⚠️  No benchmark symbols configured")
            return pd.DataFrame()

        self.logger.info(f"\n📊 Calculating benchmark returns for {len(benchmark_symbols)} symbols...")

        # ✅ ETF 가격 데이터 로드 (별도 파일)
        # 주식 데이터(price_table)와 완전히 분리하여 ETF 오염 방지
        from src.backtest.etf_data_loader import ETFDataLoader

        try:
            etf_loader = ETFDataLoader(
                config=self.config,
                root_path=self.main_ctx.root_path,
                logger=self.logger
            )

            etf_price_table = etf_loader.load_etf_prices(
                symbols=benchmark_symbols,
                start_date=start_date,
                end_date=end_date
            )

            # 주식 + ETF 통합 (벤치마크 계산용만)
            # 주의: 이 통합 테이블은 벤치마크 계산에만 사용되며,
            #      ML 학습 데이터에는 절대 사용되지 않음
            if not etf_price_table.empty:
                combined_price_table = pd.concat([
                    price_table,
                    etf_price_table
                ], ignore_index=True)
                self.logger.info(f"   Combined: {len(price_table)} stock prices + {len(etf_price_table)} ETF prices")
            else:
                self.logger.warning("   ⚠️  No ETF data loaded, using stock data only")
                combined_price_table = price_table

        except Exception as e:
            self.logger.error(f"❌ ETF data loading failed: {e}")
            self.logger.warning("   Falling back to stock price table only")
            combined_price_table = price_table

        results = []

        # 실제 거래일 찾기 (통합 테이블 사용)
        actual_start = DataProcessor.get_trade_date(pd.Timestamp(start_date), combined_price_table)
        actual_end = DataProcessor.get_trade_date(pd.Timestamp(end_date), combined_price_table)

        if actual_start is None or actual_end is None:
            self.logger.error(f"❌ Cannot find trading dates for benchmark period")
            return pd.DataFrame()

        for symbol in benchmark_symbols:
            try:
                # 심볼 가격 데이터 가져오기 (통합 테이블에서)
                # 주식이면 price_table에서, ETF면 etf_price_table에서 자동으로 찾음
                symbol_prices = combined_price_table[combined_price_table['symbol'] == symbol]

                if symbol_prices.empty:
                    self.logger.warning(f"   ⚠️  {symbol}: No price data found (symbol may not exist or typo)")
                    continue

                # 시작 가격
                start_prices = symbol_prices[symbol_prices['date'] == actual_start]
                if start_prices.empty:
                    self.logger.warning(f"   ⚠️  {symbol}: No price at start date {actual_start.date()}")
                    continue
                start_price = start_prices.iloc[0]['close']

                # 종료 가격
                end_prices = symbol_prices[symbol_prices['date'] == actual_end]
                if end_prices.empty:
                    self.logger.warning(f"   ⚠️  {symbol}: No price at end date {actual_end.date()}")
                    continue
                end_price = end_prices.iloc[0]['close']

                # 총 수익률
                total_return = (end_price - start_price) / start_price

                # 기간 내 모든 가격 데이터 (MDD, Sharpe 계산용)
                period_prices = symbol_prices[
                    (symbol_prices['date'] >= actual_start) &
                    (symbol_prices['date'] <= actual_end)
                ].sort_values('date')

                if len(period_prices) < 2:
                    self.logger.warning(f"   ⚠️  {symbol}: Insufficient price data in period")
                    continue

                # 일별 수익률
                period_prices = period_prices.copy()
                period_prices['daily_return'] = period_prices['close'].pct_change()

                # Sharpe Ratio (연율화: √252)
                daily_returns = period_prices['daily_return'].dropna()
                if len(daily_returns) > 0:
                    sharpe = (daily_returns.mean() / daily_returns.std()) * np.sqrt(252) if daily_returns.std() > 0 else 0.0
                else:
                    sharpe = 0.0

                # Maximum Drawdown
                cumulative = (1 + period_prices['daily_return'].fillna(0)).cumprod()
                running_max = cumulative.expanding().max()
                drawdown = (cumulative - running_max) / running_max
                max_drawdown = drawdown.min()

                # Win Rate (상승일 비율)
                win_rate = (daily_returns > 0).sum() / len(daily_returns) if len(daily_returns) > 0 else 0.0

                # 결과 저장
                results.append({
                    'strategy': symbol,
                    'total_return': total_return,
                    'total_return_pct': total_return * 100,
                    'sharpe_ratio': sharpe,
                    'max_drawdown': max_drawdown,
                    'max_drawdown_pct': max_drawdown * 100,
                    'win_rate': win_rate,
                    'win_rate_pct': win_rate * 100,
                    'start_price': start_price,
                    'end_price': end_price,
                    'num_days': len(period_prices)
                })

                self.logger.info(
                    f"   ✅ {symbol}: {total_return*100:+.2f}% | "
                    f"Sharpe: {sharpe:.2f} | MDD: {max_drawdown*100:.2f}%"
                )

            except Exception as e:
                self.logger.warning(f"   ⚠️  {symbol}: Calculation failed - {str(e)}")
                continue

        if not results:
            self.logger.warning("⚠️  No valid benchmark results calculated")
            return pd.DataFrame()

        return pd.DataFrame(results)

    def run(self) -> pd.DataFrame:
        """
        Walk-Forward 백테스트 실행

        Returns:
        -------
        pd.DataFrame
            백테스트 결과
        """
        self.logger.info("="*80)
        self.logger.info("ML Walk-Forward Backtest Starting (cache-based)")
        self.logger.info("="*80)
        self.logger.info(f"Rebalance period: {self.rebalance_period} months")
        self.logger.info(f"Top K: {self.top_k}")
        self.logger.info(f"Cache periods: {len(self.predictions_cache)}")

        # 가격 데이터 로드 (수익률 계산용)
        price_table = pd.read_parquet(self.main_ctx.root_path + "/processed/views/price.parquet")
        price_table['date'] = pd.to_datetime(price_table['date'])

        # Step 1: 리밸런싱 날짜 생성
        rebalance_dates = self._generate_rebalance_dates()

        # Step 2: 거래일 조정
        date_pairs = self._adjust_to_trading_days(rebalance_dates, price_table)

        # Step 3: Walk-Forward 백테스트 실행
        self._execute_walk_forward(date_pairs, price_table)

        # Step 4: 최종 리포트 및 벤치마크
        results_df, benchmark_df = self._compile_results_and_benchmark(date_pairs, price_table)

        # Step 5: Excel 레포트 저장
        self._save_backtest_report(results_df, benchmark_df, date_pairs, price_table)

        return results_df

    def _generate_rebalance_dates(self) -> list:
        """리밸런싱 날짜 목록 생성 (Multi-period 및 Single-period 모드 지원)"""
        # 우선순위: EVALUATION > BACKTEST (하위 호환성)
        eval_config = self.config.get('EVALUATION', {})
        backtest_config = self.config.get('BACKTEST', {})

        # PERIODS 읽기 (EVALUATION 우선, BACKTEST 폴백)
        periods = eval_config.get('PERIODS', backtest_config.get('PERIODS', []))

        if periods:
            # 여러 구간 모드
            self.logger.info(f"📅 Multiple backtest periods configured: {len(periods)} periods")
            source = "EVALUATION" if eval_config.get('PERIODS') else "BACKTEST"
            self.logger.info(f"   Config source: {source}.PERIODS")
            all_rebalance_dates = []

            for i, period in enumerate(periods):
                start_year = period.get('START_YEAR')
                end_year = period.get('END_YEAR')
                # 우선순위:
                # 1) PERIODS[*].START_MONTH/START_DATE
                # 2) BACKTEST.START_MONTH/START_DATE
                # 3) 최상위 START_MONTH/START_DATE (레거시 호환 앵커)
                # 4) 기본값 3/13
                start_month = int(period.get('START_MONTH', backtest_config.get('START_MONTH', self.config.get('START_MONTH', 3))))
                start_date_day = int(period.get('START_DATE', backtest_config.get('START_DATE', self.config.get('START_DATE', 13))))

                if not start_year or not end_year:
                    raise ValueError(f"Period {i+1} missing START_YEAR or END_YEAR")

                if isinstance(start_year, str):
                    start_year = int(start_year)
                if isinstance(end_year, str):
                    end_year = int(end_year)

                self.logger.info(f"  Period {i+1}: {start_year}/{start_month}/{start_date_day} ~ {end_year}/12/31")

                start_date = datetime(start_year, start_month, start_date_day)
                end_date = datetime(end_year, 12, 31)

                # 이 구간의 리밸런싱 날짜 생성
                current = start_date
                while current <= end_date:
                    all_rebalance_dates.append(current)
                    current += relativedelta(months=self.rebalance_period)

            return sorted(all_rebalance_dates)

        else:
            # 단일 구간 모드 (하위 호환성)
            start_year = backtest_config.get('START_YEAR')
            end_year = backtest_config.get('END_YEAR')

            if not start_year or not end_year:
                raise ValueError(
                    "Backtest period configuration not found!\n\n"
                    "Please configure in one of these ways:\n"
                    "  Option 1 (Recommended): Use EVALUATION section\n"
                    "    EVALUATION:\n"
                    "      PERIODS:\n"
                    "        - START_YEAR: 2020\n"
                    "          END_YEAR: 2023\n"
                    "\n"
                    "  Option 2 (Legacy): Use BACKTEST section\n"
                    "    BACKTEST:\n"
                    "      START_YEAR: 2020\n"
                    "      END_YEAR: 2023\n"
                    "\n"
                    "See config/conf.yaml.template for examples."
                )

            if isinstance(start_year, str):
                start_year = int(start_year)
            if isinstance(end_year, str):
                end_year = int(end_year)

            start_month = int(backtest_config.get('START_MONTH', self.config.get('START_MONTH', 3)))
            start_date_day = int(backtest_config.get('START_DATE', self.config.get('START_DATE', 13)))

            self.logger.info(f"📅 Single backtest period: {start_year}/{start_month}/{start_date_day} ~ {end_year}/12/31")
            self.logger.info(f"   Config source: BACKTEST (legacy mode)")

            start_date = datetime(start_year, start_month, start_date_day)
            end_date = datetime(end_year, 12, 31)

            rebalance_dates = []
            current = start_date
            while current <= end_date:
                rebalance_dates.append(current)
                current += relativedelta(months=self.rebalance_period)

            return rebalance_dates

    def _adjust_to_trading_days(self, rebalance_dates: list, price_table: pd.DataFrame) -> list:
        """리밸런싱 날짜를 실제 거래일로 조정하여 (원래날짜, 조정날짜) 튜플 리스트 반환"""
        # ✅ 거래일 조정 (regressor.py/make_mldata.py와 일원화)
        # 휴장일(주말, 공휴일)을 실제 거래 가능일로 조정
        self.logger.info(f"\n📅 Adjusting rebalance dates to actual trading days...")
        # (원래 날짜, 조정된 날짜) 튜플 리스트로 저장
        date_pairs = []

        for i, target_date in enumerate(rebalance_dates):
            # DataProcessor.get_trade_date()로 월초/월말 구분하여 거래일 찾기
            # - 월초(day <= 15): 미래 방향 (같은 분기 유지)
            # - 월말(day > 15): 과거 방향 (같은 분기 유지)
            actual_trade_date = DataProcessor.get_trade_date(pd.Timestamp(target_date), price_table)

            if actual_trade_date is None:
                self.logger.warning(
                    f"   ⚠️  Skipping {target_date.date()} - no trading day found within 10 days"
                )
                continue

            # 조정된 날짜가 원래 날짜와 다르면 로깅
            if actual_trade_date.date() != target_date.date():
                self.logger.info(
                    f"   {target_date.date()} → {actual_trade_date.date()} "
                    f"(adjusted to nearest trading day)"
                )
            else:
                self.logger.info(f"   {target_date.date()} (already a trading day)")

            # (원래 날짜, 조정된 날짜) 튜플로 저장
            original_date = target_date if isinstance(target_date, datetime) else target_date.to_pydatetime()
            adjusted_date = actual_trade_date.to_pydatetime()
            date_pairs.append((original_date, adjusted_date))

        self.logger.info(f"\n📅 Rebalance dates after adjustment: {len(date_pairs)}")
        for orig, adj in date_pairs:
            if orig.date() != adj.date():
                self.logger.info(f"   {orig.date()} → {adj.date()}")
            else:
                self.logger.info(f"   {orig.date()}")

        return date_pairs

    def _execute_walk_forward(self, date_pairs: list, price_table: pd.DataFrame):
        """Walk-Forward 백테스트 루프 실행 (캐시 기반, 결과를 self.backtest_results, self.detailed_results에 저장)"""

        for i, (original_rebalance_date, actual_rebalance_date) in enumerate(date_pairs):
            self.logger.info(f"\n{'='*80}")
            self.logger.info(f"Rebalance #{i+1}: {original_rebalance_date.date()} (actual: {actual_rebalance_date.date()})")
            self.logger.info(f"{'='*80}")

            # 1. 캐시에서 예측 로드 (regressor.py는 원본 cutoff_date를 키로 사용)
            cache_key = original_rebalance_date.strftime('%Y-%m-%d')

            if cache_key not in self.predictions_cache:
                self.logger.warning(f"⚠️ Cache miss for {cache_key}, skipping this period")
                continue

            cached_data = self.predictions_cache[cache_key]
            predictions = cached_data['predictions_df']

            if predictions is None or predictions.empty:
                self.logger.warning(f"⚠️ Cache exists but predictions_df is empty for {cache_key}, skipping")
                continue

            self.logger.info(f"📦 Using cached predictions from regressor.py")
            self.logger.info(f"   Loaded {len(predictions)} predictions from cache")
            self.logger.info(f"   Top-K selected: {len(cached_data['top_k_selected'])} stocks")

            # 2. 상위 K개 선택 (symbol + sector 포함)
            selected_stocks = self._select_top_k(predictions)
            self.logger.info(f"📊 Selected {len(selected_stocks)} stocks")

            # 3. 수익률 계산 (다음 리밸런싱 날짜까지)
            if i < len(date_pairs) - 1:
                # 다음 리밸런싱의 실제 거래일 사용
                next_actual_rebalance = date_pairs[i + 1][1]
                period_result = self._calculate_period_return(
                    selected_stocks,  # ✅ DataFrame (symbol, sector 포함)
                    actual_rebalance_date,  # 실제 매수일
                    next_actual_rebalance,   # 실제 매도일
                    price_table
                )

                avg_return = period_result['avg_return']
                self.logger.info(f"💰 Period return: {avg_return*100:.2f}%")

                # 결과 저장 (요약) - 원래 날짜와 실제 거래일 구분
                self.backtest_results.append({
                    'rebalance_date': original_rebalance_date,  # 원래 리밸런싱 날짜
                    'actual_buy_date': period_result['actual_buy_date'],
                    'actual_sell_date': period_result['actual_sell_date'],
                    'num_stocks': len(selected_stocks),  # ✅ DataFrame 길이
                    'avg_return': avg_return,
                    'retrained': False
                })

                # 상세 정보 저장 (각 종목별 + 섹터 + 카테고리 + 상장폐지 여부)
                for detail in period_result['details']:
                    self.detailed_results.append({
                        'rebalance_date': original_rebalance_date,  # 원래 리밸런싱 날짜
                        'actual_buy_date': period_result['actual_buy_date'],
                        'actual_sell_date': period_result['actual_sell_date'],
                        'symbol': detail['symbol'],
                        'sector': detail.get('sector', 'Unknown'),      # ✅ 섹터 정보 추가
                        'category': detail.get('category', 'Unknown'),  # ✅ 카테고리 정보 추가
                        'buy_price': detail['buy_price'],
                        'sell_price': detail['sell_price'],
                        'return': detail['return'],
                        'return_pct': detail['return_pct'],
                        'delisted': detail.get('delisted', False),              # 상장폐지 여부
                        'delisted_status': detail.get('delisted_status', 'active')  # ✅ 상세 상태
                    })

    def _compile_results_and_benchmark(self, date_pairs: list,
                                        price_table: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """백테스트 결과 집계 및 벤치마크 계산"""
        # 7. 최종 리포트
        results_df = pd.DataFrame(self.backtest_results)
        if not results_df.empty:
            self._print_summary(results_df)
        else:
            self.logger.warning("⚠️ No backtest results to summarize (all periods skipped or cache miss)")

        # 8. 벤치마크 계산 (전체 백테스트 기간)
        benchmark_df = pd.DataFrame()
        if len(date_pairs) > 0:
            backtest_start = date_pairs[0][1]  # 첫 번째 실제 거래일
            backtest_end = date_pairs[-1][1]   # 마지막 실제 거래일
            benchmark_df = self._calculate_benchmark_returns(backtest_start, backtest_end, price_table)

            # ML 모델 성능 추가 (비교용)
            if not results_df.empty:
                total_return = (1 + results_df['avg_return']).prod() - 1
                avg_return = results_df['avg_return'].mean()
                std_return = results_df['avg_return'].std()
                sharpe = (avg_return / std_return) * np.sqrt(4) if std_return > 0 else 0.0  # Quarterly → Annual

                # MDD 계산
                cumulative_returns = (1 + results_df['avg_return']).cumprod()
                running_max = cumulative_returns.expanding().max()
                drawdown = (cumulative_returns - running_max) / running_max
                max_drawdown = drawdown.min()

                # Win Rate
                win_rate = (results_df['avg_return'] > 0).sum() / len(results_df)

                ml_model_result = pd.DataFrame([{
                    'strategy': 'ML Model',
                    'total_return': total_return,
                    'total_return_pct': total_return * 100,
                    'sharpe_ratio': sharpe,
                    'max_drawdown': max_drawdown,
                    'max_drawdown_pct': max_drawdown * 100,
                    'win_rate': win_rate,
                    'win_rate_pct': win_rate * 100,
                    'start_price': None,
                    'end_price': None,
                    'num_days': len(results_df)
                }])

                # ML Model을 첫 번째 행으로 추가
                if not benchmark_df.empty:
                    benchmark_df = pd.concat([ml_model_result, benchmark_df], ignore_index=True)
                else:
                    benchmark_df = ml_model_result

        return results_df, benchmark_df

    def _save_backtest_report(self, results_df: pd.DataFrame, benchmark_df: pd.DataFrame,
                               date_pairs: list, price_table: pd.DataFrame):
        """백테스트 결과를 Excel 통합 레포트로 저장"""
        # 결과 저장 - Excel 통합 레포트
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_dir = Path('outputs/reports')
        report_dir.mkdir(parents=True, exist_ok=True)

        excel_file = report_dir / f'ml_backtest_report_{timestamp}.xlsx'
        detailed_df = pd.DataFrame(self.detailed_results)

        # Excel Writer로 여러 시트 저장
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Sheet 1: Summary (요약)
            results_df.to_excel(writer, sheet_name='Summary', index=False)

            # Sheet 2: Detailed (상세)
            if not detailed_df.empty:
                detailed_df.to_excel(writer, sheet_name='Detailed', index=False)

            # Sheet 3: Benchmark (벤치마크 비교) - 전체 기간 요약
            if not benchmark_df.empty:
                benchmark_df.to_excel(writer, sheet_name='Benchmark', index=False)

            # ✅ Task #7: Sheet 4 - Period-by-Period Benchmark Comparison
            if not benchmark_df.empty and not results_df.empty:
                period_comparison_df = self._create_period_benchmark_comparison(
                    results_df, date_pairs, price_table
                )
                if not period_comparison_df.empty:
                    period_comparison_df.to_excel(writer, sheet_name='Benchmark_Periods', index=False)

            # ✅ Task #6: Auto-adjust column widths for all sheets
            self._adjust_excel_column_widths(writer)

        self.logger.info(f"\n✅ Backtest report saved: {excel_file}")
        if not detailed_df.empty:
            self.logger.info(f"   Total trades: {len(detailed_df)}")
        if not benchmark_df.empty:
            self.logger.info(f"   Benchmark comparisons: {len(benchmark_df)}")


    def _print_summary(self, results: pd.DataFrame):
        """백테스트 요약 출력"""
        self.logger.info("\n" + "="*80)
        self.logger.info("BACKTEST SUMMARY")
        self.logger.info("="*80)

        total_return = (1 + results['avg_return']).prod() - 1
        avg_return = results['avg_return'].mean()
        std_return = results['avg_return'].std()
        sharpe = avg_return / std_return * np.sqrt(12/self.rebalance_period) if std_return > 0 else 0

        # MDD 계산
        cumulative = (1 + results['avg_return']).cumprod()
        running_max = cumulative.cummax()
        drawdown = (cumulative - running_max) / running_max
        mdd = drawdown.min()

        win_rate = (results['avg_return'] > 0).sum() / len(results)

        self.logger.info(f"Total Periods: {len(results)}")
        self.logger.info(f"Total Return: {total_return*100:.2f}%")
        self.logger.info(f"Average Return: {avg_return*100:.2f}%")
        self.logger.info(f"Std Dev: {std_return*100:.2f}%")
        self.logger.info(f"Sharpe Ratio: {sharpe:.2f}")
        self.logger.info(f"Max Drawdown: {mdd*100:.2f}%")
        self.logger.info(f"Win Rate: {win_rate*100:.1f}%")
        self.logger.info(f"Models Retrained: {results['retrained'].sum()} times")

    def _adjust_excel_column_widths(self, writer: pd.ExcelWriter):
        """
        ✅ Task #6: Auto-adjust Excel column widths for all sheets.
        Date columns get minimum width of 12 to display YYYY-MM-DD properly.
        """
        from openpyxl.utils import get_column_letter

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col_idx, column_cells in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                for cell in column_cells:
                    try:
                        cell_value = str(cell.value) if cell.value else ""
                        max_length = max(max_length, len(cell_value))
                    except (TypeError, ValueError):
                        pass

                # Date columns need minimum width of 12 (YYYY-MM-DD)
                header_cell = ws.cell(row=1, column=col_idx)
                header_value = str(header_cell.value) if header_cell.value else ""
                if 'date' in header_value.lower():
                    max_length = max(max_length, 12)

                # Add padding and set width
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width

    def _create_period_benchmark_comparison(
        self,
        results_df: pd.DataFrame,
        date_pairs: list,
        price_table: pd.DataFrame
    ) -> pd.DataFrame:
        """
        ✅ Task #7: Create period-by-period comparison with benchmarks.

        Returns a DataFrame with ML model returns vs benchmark returns for each period.
        """
        from src.backtest.etf_data_loader import ETFDataLoader

        benchmark_config = self.config.get('BENCHMARK', {})
        if not benchmark_config.get('ENABLED', 'N') == 'Y':
            return pd.DataFrame()

        benchmark_symbols = benchmark_config.get('SYMBOLS', [])
        if not benchmark_symbols:
            return pd.DataFrame()

        # Load ETF data
        try:
            etf_loader = ETFDataLoader(
                config=self.config,
                root_path=self.main_ctx.root_path,
                logger=self.logger
            )

            backtest_start = date_pairs[0][1]
            backtest_end = date_pairs[-1][1]

            etf_price_table = etf_loader.load_etf_prices(
                symbols=benchmark_symbols,
                start_date=backtest_start,
                end_date=backtest_end
            )

            if etf_price_table is None or etf_price_table.empty:
                self.logger.warning("⚠️ Could not load ETF prices for period comparison")
                return pd.DataFrame()

        except Exception as e:
            self.logger.warning(f"⚠️ Error loading ETF data: {e}")
            return pd.DataFrame()

        # Calculate period-by-period returns
        period_data = []

        for i, (original_date, actual_date) in enumerate(date_pairs[:-1]):
            next_actual_date = date_pairs[i + 1][1]

            # Get ML model return for this period
            ml_return = results_df.iloc[i]['avg_return'] if i < len(results_df) else 0.0

            period_row = {
                'period': i + 1,
                'start_date': actual_date,
                'end_date': next_actual_date,
                'ml_model_return': ml_return,
                'ml_model_return_pct': ml_return * 100
            }

            # Calculate benchmark returns for each symbol
            for symbol in benchmark_symbols:
                symbol_prices = etf_price_table[etf_price_table['symbol'] == symbol]

                start_price_rows = symbol_prices[symbol_prices['date'] <= actual_date].tail(1)
                end_price_rows = symbol_prices[symbol_prices['date'] <= next_actual_date].tail(1)

                if not start_price_rows.empty and not end_price_rows.empty:
                    start_price = start_price_rows.iloc[0]['close']
                    end_price = end_price_rows.iloc[0]['close']
                    bench_return = (end_price - start_price) / start_price
                else:
                    bench_return = 0.0

                period_row[f'{symbol}_return'] = bench_return
                period_row[f'{symbol}_return_pct'] = bench_return * 100

            period_data.append(period_row)

        return pd.DataFrame(period_data)
