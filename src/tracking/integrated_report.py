"""
통합 Excel 레포트 작성기

이 모듈은 다음을 결합한 통합 레포트 생성 시스템을 제공합니다:
- Regressor 예측 및 정확도 메트릭
- 백테스트 실제 거래 결과
- 벤치마크 비교

출력: 5개 시트로 구성된 단일 Excel 파일
- Sheet 1: Summary (전체 성과)
- Sheet 2: Regressor Metrics (예측 정확도)
- Sheet 3: Backtest Performance (실제 수익률)
- Sheet 4: Detailed Trades (종목별 상세 내역)
- Sheet 5: Benchmark Comparison (SPY, QQQ 등과 비교)

작성자: Quant Trading Team
날짜: 2025-12-21
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import logging


class IntegratedReportWriter:
    """
    ML 백테스트 결과를 위한 통합 Excel 레포트 작성기입니다.

    이 클래스는 regressor 예측과 백테스트 성과를 결합한 종합 Excel 레포트의
    생성을 관리합니다.

    사용 예시:
    ------
    # 작성기 생성
    writer = IntegratedReportWriter()

    # regressor 데이터 추가
    writer.add_sheet('Regressor Metrics', regressor_metrics_df)

    # ml_backtest 데이터 추가
    writer.add_sheet('Backtest Performance', backtest_results_df)

    # 파일로 저장
    filepath = writer.write()
    """

    def __init__(self, output_dir: str = "outputs/reports"):
        """
        레포트 작성기를 초기화합니다.

        Args:
            output_dir (str): 레포트를 저장할 디렉토리 (기본값: outputs/reports)
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # 파일명에 사용할 타임스탬프
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = f"integrated_report_{self.timestamp}.xlsx"
        self.filepath = self.output_dir / self.filename

        # 각 시트별 데이터 저장소
        self.sheets = {
            'Summary': None,
            'Regressor Metrics': None,
            'Backtest Performance': None,
            'Detailed Trades': None,
            'Benchmark Comparison': None
        }

        self.logger = logging.getLogger('IntegratedReport')

    def add_sheet(self, sheet_name: str, data: pd.DataFrame):
        """
        특정 시트에 데이터를 추가합니다.

        Args:
            sheet_name (str): 시트 이름 (사전 정의된 시트 중 하나여야 합니다)
            data (pd.DataFrame): 시트에 추가할 데이터

        Raises:
            ValueError: sheet_name이 인식되지 않는 경우
        """
        if sheet_name not in self.sheets:
            raise ValueError(
                f"Invalid sheet name: {sheet_name}. "
                f"Valid names: {list(self.sheets.keys())}"
            )

        if data is not None and not data.empty:
            self.sheets[sheet_name] = data
            self.logger.info(f"✅ Added sheet: {sheet_name} ({len(data)} rows)")
        else:
            self.logger.warning(f"⚠️  Sheet {sheet_name} is empty, skipping")

    def write(self) -> Path:
        """
        모든 시트를 Excel 파일로 작성합니다.

        Returns:
            Path: 생성된 Excel 파일 경로
        """
        # 비어있지 않은 시트가 하나 이상 있는지 확인
        non_empty_sheets = {k: v for k, v in self.sheets.items() if v is not None}

        if not non_empty_sheets:
            self.logger.error("❌ No data to write! All sheets are empty.")
            raise ValueError("Cannot write report: no data provided")

        self.logger.info(f"\n📊 Writing integrated report with {len(non_empty_sheets)} sheets...")

        # Excel로 작성
        with pd.ExcelWriter(self.filepath, engine='openpyxl') as writer:
            for sheet_name, df in self.sheets.items():
                if df is not None and not df.empty:
                    # 데이터 작성
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # 서식 적용
                    worksheet = writer.sheets[sheet_name]
                    self._format_sheet(worksheet, df, sheet_name)

                    self.logger.info(f"   ✅ {sheet_name}: {len(df)} rows × {len(df.columns)} cols")

        self.logger.info(f"\n✅ Integrated report saved: {self.filepath}")
        return self.filepath

    def _format_sheet(self, worksheet, df: pd.DataFrame, sheet_name: str):
        """
        Excel 시트에 서식을 적용합니다.

        Args:
            worksheet (openpyxl.worksheet.worksheet.Worksheet): Excel 워크시트 객체
            df (pd.DataFrame): 시트에 작성된 DataFrame
            sheet_name (str): 시트 이름 (컨텍스트별 서식 적용에 사용)
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # 헤더 서식
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # 헤더 스타일 적용
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 열 너비 자동 조정
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, ValueError):
                    pass

            # 너비 설정 (최대 50자)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # 헤더 행 고정
        worksheet.freeze_panes = 'A2'

        # 테두리 추가
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border


def create_summary_sheet(
    regressor_metrics: pd.DataFrame,
    backtest_performance: pd.DataFrame
) -> pd.DataFrame:
    """
    Sheet 1: Summary를 생성합니다.

    이 시트는 다음의 고수준 개요를 제공합니다:
    - 기간별 총 수익률
    - 전체 성과 메트릭
    - 기간 간 비교

    Args:
        regressor_metrics (pd.DataFrame): Regressor 평가 메트릭
        backtest_performance (pd.DataFrame): ml_backtest의 성과 결과

    Returns:
        pd.DataFrame: 요약 통계
    """
    summary = []

    if backtest_performance.empty:
        return pd.DataFrame()

    # 'period' 열이 있으면 기간별로 그룹화
    if 'period' in backtest_performance.columns:
        periods = backtest_performance['period'].unique()
    else:
        periods = ['Overall']
        backtest_performance['period'] = 'Overall'

    # 각 기간별 메트릭 계산
    for period in periods:
        period_data = backtest_performance[backtest_performance['period'] == period]

        if 'avg_return' in period_data.columns:
            returns = period_data['avg_return']
        elif 'period_return' in period_data.columns:
            returns = period_data['period_return']
        else:
            continue

        total_return = (1 + returns).prod() - 1
        avg_return = returns.mean()
        std_return = returns.std()

        # Sharpe ratio (분기별 수익률 가정)
        sharpe = (avg_return / std_return) * np.sqrt(4) if std_return > 0 else 0.0

        # 최대 낙폭
        cumulative = (1 + returns).cumprod()
        running_max = cumulative.expanding().max()
        drawdown = (cumulative - running_max) / running_max
        max_drawdown = drawdown.min()

        # 승률
        win_rate = (returns > 0).sum() / len(returns) if len(returns) > 0 else 0.0

        summary.append({
            'Period': period,
            'Total Return': f"{total_return*100:.2f}%",
            'Avg Period Return': f"{avg_return*100:.2f}%",
            'Std Dev': f"{std_return*100:.2f}%",
            'Sharpe Ratio': f"{sharpe:.2f}",
            'Max Drawdown': f"{max_drawdown*100:.2f}%",
            'Win Rate': f"{win_rate*100:.1f}%",
            'Num Periods': len(returns)
        })

    return pd.DataFrame(summary)


def create_regressor_metrics_sheet(predictions_history: List[Dict]) -> pd.DataFrame:
    """
    Sheet 2: Regressor Metrics를 생성합니다.

    이 시트는 각 리밸런싱 기간별 예측 정확도 메트릭을 표시합니다:
    - RMSE, MAE, R² (회귀 메트릭)
    - Accuracy, Precision, Recall (분류 메트릭)

    Args:
        predictions_history (List[Dict]): 기간별 예측 결과 리스트.
            각 dict는 rebalance_date, actual_returns, predicted_returns 등을 포함해야 합니다.

    Returns:
        pd.DataFrame: 기간별 Regressor 성능 메트릭
    """
    if not predictions_history:
        return pd.DataFrame()

    from sklearn.metrics import (
        mean_squared_error,
        mean_absolute_error,
        r2_score,
        accuracy_score,
        precision_score,
        recall_score
    )

    metrics = []

    for pred_info in predictions_history:
        try:
            date = pred_info['rebalance_date']
            y_true = pred_info['actual_returns']
            y_pred = pred_info['predicted_returns']

            # 이진 레이블 존재 여부 확인
            has_binary = ('actual_labels' in pred_info and
                         'predicted_labels' in pred_info)

            if has_binary:
                y_true_binary = pred_info['actual_labels']
                y_pred_binary = pred_info['predicted_labels']

            # 회귀 메트릭
            rmse = mean_squared_error(y_true, y_pred, squared=False)
            mae = mean_absolute_error(y_true, y_pred)
            r2 = r2_score(y_true, y_pred)

            # 분류 메트릭 (가능한 경우)
            if has_binary:
                accuracy = accuracy_score(y_true_binary, y_pred_binary)
                precision = precision_score(y_true_binary, y_pred_binary, zero_division=0)
                recall = recall_score(y_true_binary, y_pred_binary, zero_division=0)
            else:
                # 대체: 수익률의 부호 사용
                accuracy = accuracy_score((y_true > 0).astype(int), (y_pred > 0).astype(int))
                precision = precision_score((y_true > 0).astype(int), (y_pred > 0).astype(int), zero_division=0)
                recall = recall_score((y_true > 0).astype(int), (y_pred > 0).astype(int), zero_division=0)

            metrics.append({
                'Period': date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date),
                'RMSE': f"{rmse:.4f}",
                'MAE': f"{mae:.4f}",
                'R²': f"{r2:.4f}",
                'Accuracy': f"{accuracy*100:.2f}%",
                'Precision': f"{precision*100:.2f}%",
                'Recall': f"{recall*100:.2f}%",
                'Num Predictions': len(y_true)
            })
        except Exception as e:
            logging.warning(f"⚠️  Failed to calculate metrics for {pred_info.get('rebalance_date')}: {e}")
            continue

    return pd.DataFrame(metrics)


def create_backtest_performance_sheet(backtest_results: pd.DataFrame) -> pd.DataFrame:
    """
    Sheet 3: Backtest Performance를 생성합니다.

    이 시트는 각 리밸런싱 기간별 실제 거래 결과를 표시합니다:
    - 기간별 수익률
    - 누적 수익률
    - 거래 날짜

    Args:
        backtest_results (pd.DataFrame): ml_backtest.run()의 백테스트 결과

    Returns:
        pd.DataFrame: 서식이 적용된 백테스트 성과 데이터
    """
    if backtest_results.empty:
        return pd.DataFrame()

    # 표시용 서식 적용
    df = backtest_results.copy()

    # 날짜 서식 적용
    for col in ['rebalance_date', 'actual_buy_date', 'actual_sell_date']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col]).dt.strftime('%Y-%m-%d')

    # 퍼센트 서식 적용
    if 'avg_return' in df.columns:
        df['Period Return'] = (df['avg_return'] * 100).apply(lambda x: f"{x:.2f}%")

        # 누적 수익률 계산
        df['Cumulative Return'] = ((1 + df['avg_return']).cumprod() - 1) * 100
        df['Cumulative Return'] = df['Cumulative Return'].apply(lambda x: f"{x:.2f}%")

    # 열 선택 및 이름 변경
    display_cols = {
        'rebalance_date': 'Rebalance Date',
        'actual_buy_date': 'Buy Date',
        'actual_sell_date': 'Sell Date',
        'num_stocks': 'Stocks Selected',
        'Period Return': 'Period Return',
        'Cumulative Return': 'Cumulative Return',
        'retrained': 'Model Retrained'
    }

    available_cols = [col for col in display_cols.keys() if col in df.columns]
    result = df[available_cols].rename(columns=display_cols)

    return result


# 공개 함수 내보내기
__all__ = [
    'IntegratedReportWriter',
    'create_summary_sheet',
    'create_regressor_metrics_sheet',
    'create_backtest_performance_sheet'
]
